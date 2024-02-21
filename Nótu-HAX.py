import fitz  # PyMuPDF
import re
import imaplib
import email
from email.header import decode_header
import openpyxl
import os
from openpyxl.styles import PatternFill,Alignment, Font, Border, Side, NamedStyle
from datetime import date, datetime, timedelta
from openpyxl.worksheet.worksheet import MergedCell
from openpyxl.utils import get_column_letter

bold_font = Font(size = 22, bold=True, color = "FFFFFF", name="Times New Roman")
tan = PatternFill(start_color="DFBE6F", end_color="DFBE6F", fill_type="solid")  # Light tan color code
orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Light orange color code
hvitur = Font(bold=True, color="FFFFFF")
svartur = Font(bold=True, color="000000")
dark_blue = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark blue
orange_brown = PatternFill(start_color="CCC0B3", end_color="CCC0B3", fill_type="solid")  # Mix of orange and brown
lighter_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Lighter blue
purple_pink = PatternFill(start_color="800080", end_color="FFC0CB", fill_type="solid")  # Mix of purple and pink
thousand_separator_format = '#,##0'
indiced_kaup = 0

def is_company(company_name):
    listi = ["Arctica", "Fossar", "Íslensk", "ACRO", "Íslandsbanki", "Kvika", "Landsbankinn"]
    if company_name in listi:
        return True
    return False

filepath = None
all_email = []
email_address = "test_notur_@outlook.com"
password = "Stefnir1234"
imap_server = "outlook.office365.com"

mail = imaplib.IMAP4_SSL(imap_server)

mail.login(email_address, password)

mail.select("inbox")

# Search for all unseen emails
status, messages = mail.search(None, "(UNSEEN)")
# Get the list of email IDs
email_ids = messages[0].split()

def run():
    for email_id in email_ids:
        # Fetch the email by ID
        status, msg_data = mail.fetch(email_id, "(RFC822)")
        
        # Get the email content
        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)
        
        # Get the subject and sender
        subject, encoding = decode_header(msg["Subject"])[0]
        sender, encoding = decode_header(msg.get("From"))[0]

        # print("Subject:", subject)
        # print("From:", sender)
        
        if msg.is_multipart():
            for part in msg.walk():
                dic = {}
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))

                if "attachment" in content_disposition:
                    filename = decode_header(part.get_filename())[0][0]
                    print(filename)
                    if filename:
                        if isinstance(filename, bytes):
                            try:
                                filename = filename.decode("utf-8")
                            except UnicodeDecodeError:
                                try:
                                    filename = filename.decode("latin1")  # Try a different encoding
                                except UnicodeDecodeError:
                                    filename = filename.decode("utf-8", errors='replace')  
                        filepath = os.path.join("Desktop", filename) # Þetta þarf að vera breytt svo allar nóturnar vistast ehvstaðar
                        
                        # Save the attachment to a file
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                            
                        # print(f"Attachment saved: {filepath}")
                        with fitz.open(filepath) as pdf_document:
                            dic = {}
                            num_pages = pdf_document.page_count
                            array = []
                            for page_num in range(num_pages):
                                page = pdf_document[page_num]
                                text = page.get_text()
                                # Miðlari
                                isb = False
                                match = re.search(r'Miðlari:\s*(.*)', text)
                                if match:
                                    broker_name = match.group(1).strip()
                                    dic["Miðlari"] = broker_name
                                match = re.search(r'Sölumaður :\s*(.*)', text)
                                if match:
                                    isb = True
                                    broker_name = match.group(1).strip()
                                    dic["Miðlari"] = broker_name
                                texti = text.split(" ")
                                # print(texti)
                                if(not isb):
                                    for i in range(0, len(texti)):
                                        splitmac = texti[i].split("\n")
                                        for x in splitmac:
                                            if(is_company(x)):
                                                if(x == "Íslensk"):
                                                    dic["Fyrirtæki"] = "Íslensk verðbréf"
                                                else:
                                                    dic["Fyrirtæki"] = x 
                                        print(texti[i])
                                        if("Dags" in texti[i] and "viðskipta" in texti[i+1]):
                                            date_string = texti[i + 1][10:]  # Start from the 10th character
                                            # Find the position of the first occurrence of four consecutive numbers
                                            match = re.search(r'\d{4}', date_string)
                                            if match:
                                                # End the substring right at the four consecutive numbers
                                                end_index = match.start()
                                                substring = date_string[:end_index+4]
                                            dic["Dags viðskipta"] = substring
                                        if("númer" in texti[i]):
                                            nytt = texti[i].split("\n")
                                            dic["IS númer"] = nytt[1]
                                        if("Þú" in texti[i] and ("seldir" in texti[i+1] or "keyptir" in texti[i+1])):
                                            index_of_newline = texti[i+1].find('\n')
                                            if index_of_newline != -1 and index_of_newline < len(texti[i+1]) - 1:
                                                name = texti[i+1][index_of_newline + 1:]
                                                name = name.strip()
                                                if("seldir" in texti[i+1]):
                                                    dic["Seldir"] = name
                                                else:
                                                    dic["Keyptir"] = name

                                        if("Upphæð" in texti[i]):
                                            uppl = texti[i].split("\n")
                                            dic["Upphæð"] = uppl[4]
                                            dic["Gengi"] = uppl[5]
                                            dic["Hlutir"] = uppl[6]
                                        if("Þóknun" in texti[i]):
                                            thok =  texti[i+1][1:-1]
                                            dic["Þóknun"] = thok
                                else:
                                    keyp = True
                                    for i in range(0, len(texti)):
                                        splitmac = texti[i].split("\n")
                                        for x in splitmac:
                                            if(is_company(x)):
                                                if(x == "Íslensk"):
                                                    dic["Fyrirtæki"] = "Íslensk verðbréf"
                                                else:
                                                    dic["Fyrirtæki"] = x 
                                        
                                        if("Dags" in texti[i] and "viðskipta" in texti[i+1]):
                                            date_string = texti[i + 2]
                                            # Find the position of the first occurrence of four consecutive numbers
                                            match = re.search(r'\d{4}', date_string)
                                            if match:
                                                # End the substring right after the four consecutive numbers
                                                end_index = match.start() + 4
                                                substring = date_string[:end_index]
                                            else:
                                                # If no match is found, use the whole string or handle accordingly
                                                substring = date_string

                                            dic["Dags viðskipta"] = substring
                                        if("ISIN" in texti[i]):
                                            nytt = texti[i+3].split("\n")
                                            dic["IS númer"] = nytt[0]
                                        if("seldum" in texti[i]):
                                            keyp = False
                                        if("Upphæð" in texti[i]):
                                            uppl = texti[i].split("\n")
                                            dic["Gengi"] = uppl[4]
                                            dic["Upphæð"] = uppl[6]
                                            if(keyp):
                                                dic["Keyptir"] = uppl[0]
                                            else:
                                                dic["Seldir"] = uppl[0]
                                        if("Nafnverð" in texti[i]):
                                            uppl = texti[i].split("\n")
                                            dic["Hlutir"] = uppl[2]
                                        if("Þóknun" in texti[i]):
                                            uppl = texti[i].split("\n")
                                            thok = uppl[4]
                                            thok = float(thok.replace('.', ''))
                                            upph = float(dic["Upphæð"].replace(".", ""))
                                            res = int(thok) / upph
                                            dic["Þóknun"] = res
                                    
                        ### Checka ef lengdin i dictionarinu se ekki örugglega jafn mikið og upplysingarnar
                else:
                    if part.get_payload():
                        # Extract and print the email body
                        if part.get_content_type() == "text/plain":
                            body = part.get_payload(decode=True).decode("utf-8", "ignore")
                            bod = body.split("\r")
                            if("landsbank" in body):
                                landsbank = True
                                dic["Fyrirtæki"] = "Landsbankinn"
                                for i in range(0, len(bod)):
                                    print(bod[i])
                                    if("iðlari" in bod[i]):
                                        index = bod[i].find(":")
                                        dic["Miðlari"] = bod[i][index+1:].strip()
                                    if("Samtals" in bod[i]):
                                        dic["Upphæð"] = bod[i+1][1:]
                                    if("ISIN" in bod[i]):
                                        split = bod[i+1].split("\n")
                                        dic["IS númer"] = split[1]
                                    if("Þóknun" in bod[i]):
                                        index = bod[i].find("(")
                                        dic["Þóknun"] = bod[i][index + 1:-1].strip()
                                    if("Verð" in bod[i]):
                                        dic["Gengi"] = bod[i+4][1:]
                                        dic["Hlutir"] = bod[i+3][1:]
                                    if("Dags. viðskipta" in bod[i]):
                                        dic["Dags viðskipta"] = bod[i+1][1:]
                                    if("Sala" in bod[i]):
                                        dic["Seldir"] = bod[i+1][1:]
                                    if("Kaup" in bod[i]):
                                        dic["Keyptir"] = bod[i+1][1:]
                all_email.append(dic)
    mail.logout()
    isb = False
    return all_email

##### ctrl r til að runna

result = run()


def sorting_key(dictionary):
    sala = []  # Array for dictionaries with "Seldir"
    kaup = []  # Array for dictionaries with "Keyptir"

    if "Seldir" in dictionary:
        sala.append(dictionary)
    elif "Keyptir" in dictionary:
        kaup.append(dictionary)

    return ()


sorted_array = sorted(result, key=sorting_key)
sala = [d for d in sorted_array if "Seldir" in d]
kaup = [d for d in sorted_array if "Keyptir" in d]

def error_check():
    for i in sala:
        if(len(i) != 9):
            return True
    for x in kaup:
        if(len(x)!=9):
            return True
    return False 


def error(check):
    for i in check:
        if("Þóknun" not in i):
            i["Þóknun"] = "0"
        if("Miðlari" not in i):
            i["Miðlari"] = "Villa"
        if("Dags viðskipta" not in i):
            today_date = datetime.today().date()
            # Format it as "%d.%m.%Y"
            formatted_date = today_date.strftime("%d.%m.%Y")
            i["Dags viðskipta"] = formatted_date
        if("Upphæð" not in i):
            i["Upphæð"] = "0"
        if("Hlutir" not in i):
            i["Hlutir"] = "0"
        if("Gengi" not in i):
            i["Gengi"] = "0"
        if("IS númer" not in i):
            i["IS númer"] = "0"
        if("Fyrirtæki" not in i):
            i["Fyrirtæki"] = "Villa"
 
error(sala)
error(kaup)

numer_sjoða = {"Stefnir - Innlend hlutabréf": 266419, "Stefnir - ÍS 5": 435690, "Stefnir - Innlend hlutabréf Vogun hs.": 506728, "Stefnir - Arðgreiðslusjóður": 500013}
kt_sjoða = {"Stefnir - Innlend hlutabréf": "490206-8450", "Stefnir - ÍS 5": "420407-9610", "Stefnir - Innlend hlutabréf Vogun hs.": "510422-9960", "Stefnir - Arðgreiðslusjóður": "581120-9740"}
ekki_grænan_sjoða = {"Stefnir - Innlend hlutabréf": "Stefnir - 5408", "Stefnir - ÍS 5": "Eignastýring 488885", "Stefnir - Innlend hlutabréf Vogun hs.": "Stefnir - 004763", "Stefnir - Arðgreiðslusjóður": "Stefnir - 559812"}
key_array = list(ekki_grænan_sjoða.keys())

workbook = openpyxl.Workbook()
sheet = workbook.active
row_num = [num for num in range(3, 101) if num % 5 != 2]

#Raðir sem eru hvítar, þ.e.a.s. 7, 12, 17...
vantar_row = []
for i in range(3, 100):
    if(i not in row_num):
        vantar_row.append(i)
indiced = len(sala) * 4
tafla_index = 2
print(sala)
print(kaup)
## Upplýsingar####################################################################
if(len(sala) != 0):
    # dags = sala[0]["Dags viðskipta"]
    # date_object = datetime.strptime(dags, "%d.%m.%Y").date()
    # date_object_2 = date_object + timedelta(days=2)
    i = 0
    fyrirtæki = 0
    for ind in range(0, indiced):
        dags = sala[ind // 4]["Dags viðskipta"]
        try:
            date_object = datetime.strptime(dags, "%d.%m.%Y").date()
        except Exception as e:
            date_object = datetime.today().date()
            # Format it as "%d.%m.%Y"
        date_object_2 = date_object + timedelta(days=2)
        cell_b = sheet[f"B{row_num[ind]}"]
        cell_c = sheet[f"C{row_num[ind]}"]
        cell_d = sheet[f"D{row_num[ind]}"]
        cell_e = sheet[f"E{row_num[ind]}"]
        cell_b.value = "Sala"
        cell_c.value = dags
        cell_d.value = date_object_2
        cell_e.value = "00:00:00"

        cell_c.fill = tan
        cell_e.fill = tan

        key_index = ind % len(key_array)
        cell_i = sheet[f"I{row_num[ind]}"]
        cell_i.value = ekki_grænan_sjoða[key_array[key_index]]
        cell_g = sheet[f"G{row_num[ind]}"]
        cell_g.value = kt_sjoða[key_array[key_index]]
        cell_f = sheet[f"F{row_num[ind]}"]
        cell_f.value = numer_sjoða[key_array[key_index]]
        cell_h = sheet[f"H{row_num[ind]}"]
        cell_h.value = key_array[i % len(key_array)]

        cell_n = sheet[f"N{row_num[ind]}"]
        cell_n.value = 0.5
        cell_n.number_format = '0.00%'
        cell_n.fill = orange
        cell_p = sheet[f"P{row_num[ind]}"]
        cell_p.value = sala[ind // 4]["Seldir"]
        cell_p.fill = tan
        cell_t = sheet[f"T{row_num[ind]}"]
        cell_t.value = sala[ind // 4]["Gengi"]
        cell_t.fill = tan
        cell_u = sheet[f"U{row_num[ind]}"]
        cell_u.value = sala[ind // 4]["Þóknun"]
        cell_u.fill = tan
        cell_q = sheet[f"Q{row_num[ind]}"]
        cell_q.value = sala[ind // 4]["IS númer"]
        cell_q.fill = tan
        # skipting = float(input(f"Sala á {sala[ind // 4]["Seldir"]}. Hluti til {key_array[i % len(key_array)]} (sem kommutala): "))
        # cell_n.value = skipting
        skipting = 1 ##Breyta seinna
        cell_r = sheet[f"R{row_num[ind]}"]
        float_gengi = float(sala[ind // 4]["Gengi"].replace(',', '.'))
        float_hlutir = float(sala[ind // 4]["Hlutir"].replace('.', ''))

        cell_r.value = f'=+{float_hlutir}*N{row_num[ind]}'
        cell_r.fill = tan
        cell_s = sheet[f"S{row_num[ind]}"]
        cell_s.value = f'=+R{row_num[ind]}*T{row_num[ind]}'
        cell_s.fill = tan
        cell_w = sheet[f"W{row_num[ind]}"]
        cell_w.value = sala[ind // 4]["Fyrirtæki"]
        cell_w.fill = tan
        cell_x = sheet[f"X{row_num[ind]}"]
        cell_x.value = sala[ind // 4]["Miðlari"]
        cell_x.fill = tan

        cell_v = sheet[f"V{row_num[ind]}"]
        cell_v.value = f'=+S{row_num[ind]}*(1-U{row_num[ind]})'
        cell_v.fill = tan
        i += 1
        tafla_index = row_num[ind]

for row_number in range(1, 3):
    for col_number in range(2, 25):
        cell = sheet.cell(row=row_number, column=col_number)
        if(col_number in [10,11,12,13]):
            cell.fill = PatternFill(start_color="91A0AE", end_color="91A0AE", fill_type="solid")  # Light Blue
        else:
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark Blue           

if(len(kaup) != 0):
    indiced_kaup = len(kaup) * 4 + indiced
    # dags = kaup[row_num[ind]]["Dags viðskipta"]
    # date_object = datetime.strptime(dags, "%d.%m.%Y").date()
    # date_object_2 = date_object + timedelta(days=2)
    i = 0
    auka_index = 0
    for ind in range(indiced, indiced_kaup):
        dags = kaup[auka_index // 4]["Dags viðskipta"]
        try:
            date_object = datetime.strptime(dags, "%d.%m.%Y").date()
        except Exception as e:
            date_object = datetime.today().date()
        date_object_2 = date_object + timedelta(days=2)
        cell_b = sheet[f"B{row_num[ind]}"]
        cell_c = sheet[f"C{row_num[ind]}"]
        cell_d = sheet[f"D{row_num[ind]}"]
        cell_e = sheet[f"E{row_num[ind]}"]
        cell_b.value = "Kaup"
        cell_c.value = dags
        cell_d.value = date_object_2
        cell_e.value = "00:00:00"

        cell_c.fill = tan
        cell_e.fill = tan

        key_index = ind % len(key_array)
        cell_m = sheet[f"M{row_num[ind]}"]
        cell_m.value = ekki_grænan_sjoða[key_array[key_index]]
        cell_k = sheet[f"K{row_num[ind]}"]
        cell_k.value = kt_sjoða[key_array[key_index]]
        cell_j = sheet[f"J{row_num[ind]}"]
        cell_j.value = numer_sjoða[key_array[key_index]]
        cell_l = sheet[f"L{row_num[ind]}"]
        cell_l.value = key_array[i % len(key_array)]

        cell_n = sheet[f"N{row_num[ind]}"]
        cell_n.fill = orange
        cell_n.value = 0.5
        cell_n.number_format = '0.00%'
        cell_p = sheet[f"P{row_num[ind]}"]
        cell_p.value = kaup[auka_index // 4]["Keyptir"]
        cell_p.fill = tan
        cell_t = sheet[f"T{row_num[ind]}"]
        cell_t.value = kaup[auka_index // 4]["Gengi"]
        cell_t.fill = tan
        cell_u = sheet[f"U{row_num[ind]}"]
        cell_u.value = kaup[auka_index // 4]["Þóknun"]
        cell_u.fill = tan
        cell_q = sheet[f"Q{row_num[ind]}"]
        cell_q.value = kaup[auka_index // 4]["IS númer"]
        cell_q.fill = tan

        cell_r = sheet[f"R{row_num[ind]}"]
        float_gengi = float(kaup[auka_index // 4]["Gengi"].replace(',', '.'))
        float_hlutir = float(kaup[auka_index // 4]["Hlutir"].replace('.', ''))

        cell_r.value = f'=+{float_hlutir}*N{row_num[ind]}'
        cell_r.fill = tan
        cell_s = sheet[f"S{row_num[ind]}"]
        cell_s.value = f'=+R{row_num[ind]}*T{row_num[ind]}'
        cell_s.fill = tan
        cell_w = sheet[f"W{row_num[ind]}"]
        cell_w.value = kaup[auka_index // 4]["Fyrirtæki"]
        cell_w.fill = tan
        cell_x = sheet[f"X{row_num[ind]}"]
        cell_x.value = kaup[auka_index // 4]["Miðlari"]
        cell_x.fill = tan

        cell_v = sheet[f"V{row_num[ind]}"]
        cell_v.value = f'=+S{row_num[ind]}*(1+U{row_num[ind]})'
        cell_v.fill = tan
        i += 1
        auka_index += 1
        tafla_index = row_num[ind]
        print(ind)


#### Give sum to white columns, summing hlutar keyptir and total amount
if(indiced_kaup==0):
    indiced_kaup = indiced

for grey in vantar_row:
    if grey < tafla_index + 5:
        cell_r_sum = sheet[f'R{grey}']
        cell_r_sum.value =f'=SUM(R{grey - 4}:R{grey - 1})'
        sum_formula_v = f'=SUM(V{grey - 4}:V{grey - 1})'
        cell_v_sum = sheet[f'V{grey}']
        cell_v_sum.value = sum_formula_v

#################################### Tafla 1: kaup og sala ######################################################################
indiced_kaup = tafla_index + 2 #+ auka_index
cell_d_sjod = sheet[f"D{indiced_kaup+5}"]
cell_e_sjod = sheet[f"E{indiced_kaup+5}"]
cell_f_sjod = sheet[f"F{indiced_kaup+5}"]
cell_g_sjod = sheet[f"G{indiced_kaup+5}"]
cell_c_kaup = sheet[f"C{indiced_kaup+6}"]
cell_c_sala = sheet[f"C{indiced_kaup+7}"]
cell_c_net = sheet[f"C{indiced_kaup+8}"]
# kaup
cell_d_inn_kaup = sheet[f"D{indiced_kaup+6}"]
cell_e_is_kaup = sheet[f"E{indiced_kaup+6}"]
cell_f_vog_kaup = sheet[f"F{indiced_kaup+6}"]
cell_g_arð_kaup = sheet[f"G{indiced_kaup+6}"]
# Sala
cell_d_inn_sala = sheet[f"D{indiced_kaup+7}"]
cell_e_is_sala = sheet[f"E{indiced_kaup+7}"]
cell_f_vog_sala = sheet[f"F{indiced_kaup+7}"]
cell_g_arð_sala = sheet[f"G{indiced_kaup+7}"]
#Net
cell_d_inn_net = sheet[f"D{indiced_kaup+8}"]
cell_e_is_net = sheet[f"E{indiced_kaup+8}"]
cell_f_vog_net = sheet[f"F{indiced_kaup+8}"]
cell_g_arð_net = sheet[f"G{indiced_kaup+8}"]

cell_d_sjod.value = key_array[0]
cell_e_sjod.value = key_array[1]
cell_f_sjod.value = key_array[2]
cell_g_sjod.value = key_array[3]
cell_c_kaup.value = "Kaup"
cell_c_sala.value = "Sala"
cell_c_net.value = "Net"
# Kaup
cell_d_inn_kaup.value = f'=SUMIFS(S2:S{indiced_kaup+4}, J2:J{indiced_kaup+4}, ">0", L2:L{indiced_kaup+4}, D{indiced_kaup+5})'
cell_e_is_kaup.value = f'=SUMIFS(S2:S{indiced_kaup+4}, J2:J{indiced_kaup+4}, ">0", L2:L{indiced_kaup+4}, E{indiced_kaup+5})'
cell_f_vog_kaup.value = f'=SUMIFS(S2:S{indiced_kaup+4}, J2:j{indiced_kaup+4}, ">0", L2:L{indiced_kaup+4}, F{indiced_kaup+5})'
cell_g_arð_kaup.value = f'=SUMIFS(S2:S{indiced_kaup+4}, J2:J{indiced_kaup+4}, ">0", L2:L{indiced_kaup+4}, G{indiced_kaup+5})'
# Sala
cell_d_inn_sala.value = f'=SUMIFS(S2:S{indiced_kaup+4}, F2:F{indiced_kaup+4}, ">0", H2:H{indiced_kaup+4}, D{indiced_kaup+5})'
cell_e_is_sala.value = f'=SUMIFS(S2:S{indiced_kaup+4}, F2:F{indiced_kaup+4}, ">0", H2:H{indiced_kaup+4}, E{indiced_kaup+5})'
cell_f_vog_sala.value = f'=SUMIFS(S2:S{indiced_kaup+4}, F2:F{indiced_kaup+4}, ">0", H2:H{indiced_kaup+4}, F{indiced_kaup+5})'
cell_g_arð_sala.value = f'=SUMIFS(S2:S{indiced_kaup+4}, F2:F{indiced_kaup+4}, ">0", H2:H{indiced_kaup+4}, G{indiced_kaup+5})'
#Net
cell_d_inn_net.value = f'=+D{indiced_kaup+7}-D{indiced_kaup+6}'
cell_e_is_net.value = f'=+E{indiced_kaup+7}-E{indiced_kaup+6}'
cell_f_vog_net.value = f'=+F{indiced_kaup+7}-F{indiced_kaup+6}'
cell_g_arð_net.value = f'=+G{indiced_kaup+7}-G{indiced_kaup+6}'

for col_num in range(3, 8):
    row_num = indiced_kaup + 7
    cell = sheet.cell(row=row_num, column=col_num)
    border = Border(bottom=Side(style='thin'))
    cell.border = border

for cell in [
             cell_c_kaup, cell_c_sala, cell_c_net,
             cell_d_inn_kaup, cell_e_is_kaup, cell_f_vog_kaup, cell_g_arð_kaup,
             cell_d_inn_sala, cell_e_is_sala, cell_f_vog_sala, cell_g_arð_sala, cell_d_inn_net, cell_e_is_net, cell_f_vog_net, cell_g_arð_net]:
    cell.number_format = thousand_separator_format

cell_d_sjod.font = hvitur
cell_d_sjod.fill = dark_blue

cell_e_sjod.fill = orange_brown
cell_e_sjod.font = svartur

cell_f_sjod.fill = lighter_blue
cell_f_sjod.font = svartur

cell_g_sjod.font = hvitur
cell_g_sjod.fill = purple_pink

#################################### Tafla 2: kaup og sala niður á félag ######################################################################
sheet.merge_cells(f'D{indiced_kaup + 11}:E{indiced_kaup + 11}')
merged_cell_innlent = sheet[f'D{indiced_kaup + 11}']
merged_cell_innlent.value = key_array[0]
merged_cell_innlent.fill = dark_blue
merged_cell_innlent.font = hvitur
cell_d_kaup = sheet[f"D{indiced_kaup+12}"]
cell_d_kaup.value = "Kaup"
cell_e_kaup = sheet[f"E{indiced_kaup+12}"]
cell_e_kaup.value = "Sala"

sheet.merge_cells(f'F{indiced_kaup + 11}:G{indiced_kaup + 11}')
merged_cell_is_5 = sheet[f'F{indiced_kaup + 11}']
merged_cell_is_5.value = key_array[1]
merged_cell_is_5.fill = orange_brown
merged_cell_is_5.font = svartur
cell_f_kaup = sheet[f"F{indiced_kaup+12}"]
cell_f_kaup.value = "Kaup"
cell_g_kaup = sheet[f"G{indiced_kaup+12}"]
cell_g_kaup.value = "Sala"

sheet.merge_cells(f'H{indiced_kaup + 11}:I{indiced_kaup + 11}')
merged_cell_inn_vog = sheet[f'H{indiced_kaup + 11}']
merged_cell_inn_vog.value = key_array[2]
merged_cell_inn_vog.fill = lighter_blue
merged_cell_inn_vog.font = svartur
cell_h_kaup = sheet[f"H{indiced_kaup+12}"]
cell_h_kaup.value = "Kaup"
cell_i_kaup = sheet[f"I{indiced_kaup+12}"]
cell_i_kaup.value = "Sala"

sheet.merge_cells(f'J{indiced_kaup + 11}:K{indiced_kaup + 11}')
merged_cell_arður = sheet[f'J{indiced_kaup + 11}']
merged_cell_arður.value = key_array[3]
merged_cell_arður.fill = purple_pink
merged_cell_arður.font = hvitur
cell_j_kaup = sheet[f"J{indiced_kaup+12}"]
cell_j_kaup.value = "Kaup"
cell_k_kaup = sheet[f"K{indiced_kaup+12}"]
cell_k_kaup.value = "Sala"

fresh_index = indiced_kaup + 12 
distinct_company = []
for info in sala:
    distinct_company.append(info["Seldir"])
for infor in kaup:
    distinct_company.append(infor["Keyptir"])

for i in range(0, len(distinct_company)):
    cell_comp = sheet[f"C{fresh_index + i + 1}"]
    cell_comp.value = distinct_company[i]
    
    cell_comp_kaup_inn_hlut = sheet[f"D{fresh_index + i + 1}"]
    cell_comp_kaup_inn_hlut.value = f'=SUMIFS(S2:S{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, D{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_kaup_inn_hlut.number_format = thousand_separator_format

    cell_comp_sala_inn_hlut = sheet[f"E{fresh_index + i + 1}"]
    cell_comp_sala_inn_hlut.value = f'=SUMIFS(S2:S{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, D{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_sala_inn_hlut.number_format = thousand_separator_format

    cell_comp_kaup_is_5 = sheet[f"F{fresh_index + i + 1}"]
    cell_comp_kaup_is_5.value = f'=SUMIFS(S2:S{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, F{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_kaup_is_5.number_format = thousand_separator_format

    cell_comp_sala_is_5 = sheet[f"G{fresh_index + i + 1}"]
    cell_comp_sala_is_5.value = f'=SUMIFS(S2:S{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, F{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_sala_is_5.number_format = thousand_separator_format   

    cell_comp_kaup_inn_vog = sheet[f"H{fresh_index + i + 1}"]
    cell_comp_kaup_inn_vog.value = f'=SUMIFS(S2:S{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, H{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_kaup_inn_vog.number_format = thousand_separator_format   

    cell_comp_sala_inn_vog = sheet[f"I{fresh_index + i + 1}"]
    cell_comp_sala_inn_vog.value = f'=SUMIFS(S2:S{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, H{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_sala_inn_vog.number_format = thousand_separator_format  

    cell_comp_kaup_arð = sheet[f"J{fresh_index + i + 1}"]
    cell_comp_kaup_arð.value = f'=SUMIFS(S2:S{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, J{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_kaup_arð.number_format = thousand_separator_format  

    cell_comp_sala_arð = sheet[f"K{fresh_index + i + 1}"]
    cell_comp_sala_arð.value = f'=SUMIFS(S2:S{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, J{fresh_index - 1}, P2:P{fresh_index - 2}, C{fresh_index + i + 1})'
    cell_comp_sala_arð.number_format = thousand_separator_format  

    for cell in [cell_comp_kaup_inn_hlut, cell_comp_sala_inn_hlut, cell_comp_kaup_is_5, 
                cell_comp_sala_is_5, cell_comp_kaup_inn_vog, cell_comp_sala_inn_vog, cell_comp_kaup_arð, cell_comp_sala_arð]:
        cell.number_format = thousand_separator_format
####### Tafla 3: Keyptir hlutar ####################################
sheet.merge_cells(f'D{fresh_index + i + 5}:E{fresh_index + i + 5}')
merged_cell_inn = sheet[f'D{fresh_index + i + 5}']
merged_cell_inn_bonus = sheet[f'E{fresh_index + i + 5}']
merged_cell_inn.value = key_array[0]
merged_cell_inn.fill = dark_blue
merged_cell_inn.font = hvitur
merged_cell_inn_bonus.font = hvitur
cell_d_kaup = sheet[f"D{fresh_index + i + 6}"]
cell_d_kaup.value = "Kaup"
cell_e_kaup = sheet[f"E{fresh_index + i + 6}"]
cell_e_kaup.value = "Sala"

sheet.merge_cells(f'F{fresh_index + i + 5}:G{fresh_index + i + 5}')
merged_cell_is_5 = sheet[f'F{fresh_index + i + 5}']
merged_cell_is_5.value = key_array[1]
merged_cell_is_5.fill = orange_brown
merged_cell_is_5.font = svartur
cell_f_kaup = sheet[f"F{fresh_index + i + 6}"]
cell_f_kaup.value = "Kaup"
cell_g_kaup = sheet[f"G{fresh_index + i + 6}"]
cell_g_kaup.value = "Sala"

sheet.merge_cells(f'H{fresh_index + i + 5}:I{fresh_index + i + 5}')
merged_cell_inn_vog = sheet[f'H{fresh_index + i + 5}']
merged_cell_inn_vog.value = key_array[2]
merged_cell_inn_vog.fill = lighter_blue
merged_cell_inn_vog.font = svartur
cell_h_kaup = sheet[f"H{fresh_index + i + 6}"]
cell_h_kaup.value = "Kaup"
cell_i_kaup = sheet[f"I{fresh_index + i + 6}"]
cell_i_kaup.value = "Sala"

sheet.merge_cells(f'J{fresh_index + i + 5}:K{fresh_index + i + 5}')
merged_cell_arð = sheet[f'J{fresh_index + i + 5}']
merged_cell_arð_bonus = sheet[f'K{fresh_index + i + 5}']
merged_cell_arð.value = key_array[3]
merged_cell_arð.fill = purple_pink
merged_cell_arð.font = hvitur
merged_cell_arð_bonus.font = hvitur
cell_j_kaup = sheet[f"J{fresh_index + i + 6}"]
cell_j_kaup.value = "Kaup"
cell_k_kaup = sheet[f"K{fresh_index + i + 6}"]
cell_k_kaup.value = "Sala"

using = fresh_index + i + 6
for i in range(0, len(distinct_company)):
    cell_comp = sheet[f"C{using + i + 1}"]
    cell_comp.value = distinct_company[i]

    cell_comp_kaup_inn_hlut = sheet[f"D{using + i + 1}"]
    cell_comp_kaup_inn_hlut.value = f'=SUMIFS(R2:R{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, D{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_kaup_inn_hlut.number_format = thousand_separator_format

    cell_comp_sala_inn_hlut = sheet[f"E{using + i + 1}"]
    cell_comp_sala_inn_hlut.value = f'=SUMIFS(R2:R{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, D{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_sala_inn_hlut.number_format = thousand_separator_format

    cell_comp_kaup_is_5 = sheet[f"F{using + i + 1}"]
    cell_comp_kaup_is_5.value = f'=SUMIFS(R2:R{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, F{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_kaup_is_5.number_format = thousand_separator_format

    cell_comp_sala_is_5 = sheet[f"G{using + i + 1}"]
    cell_comp_sala_is_5.value = f'=SUMIFS(R2:R{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, F{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_sala_is_5.number_format = thousand_separator_format   

    cell_comp_kaup_inn_vog = sheet[f"H{using + i + 1}"]
    cell_comp_kaup_inn_vog.value = f'=SUMIFS(R2:R{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, H{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_kaup_inn_vog.number_format = thousand_separator_format   

    cell_comp_sala_inn_vog = sheet[f"I{using + i + 1}"]
    cell_comp_sala_inn_vog.value = f'=SUMIFS(R2:R{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, H{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_sala_inn_vog.number_format = thousand_separator_format  

    cell_comp_kaup_arð = sheet[f"J{using + i + 1}"]
    cell_comp_kaup_arð.value = f'=SUMIFS(R2:R{fresh_index - 2}, J2:J{fresh_index - 2}, ">0", L2:L{fresh_index - 2}, J{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_kaup_arð.number_format = thousand_separator_format  

    cell_comp_sala_arð = sheet[f"K{using + i + 1}"]
    cell_comp_sala_arð.value = f'=SUMIFS(R2:R{fresh_index - 2}, F2:F{fresh_index - 2}, ">0", H2:H{fresh_index - 2}, J{using - 1}, P2:P{fresh_index - 2}, C{using + i + 1})'
    cell_comp_sala_arð.number_format = thousand_separator_format  

    for cell in [cell_comp_kaup_inn_hlut, cell_comp_sala_inn_hlut, cell_comp_kaup_is_5, 
                cell_comp_sala_is_5, cell_comp_kaup_inn_vog, cell_comp_sala_inn_vog, cell_comp_kaup_arð, cell_comp_sala_arð]:
        cell.number_format = thousand_separator_format
#################################### Autofit every cell ########################################################################

fixed_width = 34

for col_num, column in enumerate(sheet.columns, start=1):
    max_length = 0
    column = [cell for cell in column]

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    # Determine the width based on the column number
    if 4 <= col_num <= 13:  # Columns D to M
        adjusted_width = fixed_width
    else:
        adjusted_width = (max_length + 3)  # Adjust this value for other columns

    column_letter = get_column_letter(col_num)

    # Set column width
    sheet.column_dimensions[column_letter].width = adjusted_width

    # Set font style (Times New Roman)
    for cell in column:
        cell.font = Font(name='Times New Roman')

    # Set alignment to center
    for cell in column:
        cell.alignment = Alignment(horizontal='center')

    # Add bottom border
    for row_num, cell in enumerate(column, start=1):
        if row_num <= indiced_kaup:
            cell.border = Border(bottom=Side(style='thin'))

#################################### Set headers and merge and center ########################################################################
sheet.merge_cells('F1:I2')
merged_cell = sheet['F1']
merged_cell.value = 'Seljandi'
merged_cell.font = bold_font
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
cell_d_sjod.font = hvitur
cell_g_sjod.font = hvitur
merged_cell_inn_bonus.font = hvitur
merged_cell_inn.font = hvitur
merged_cell_arð.font = hvitur
merged_cell_arður.font = hvitur
merged_cell_innlent.font = hvitur

sheet.merge_cells('J1:M2')
merged_cell2 = sheet['J1']
merged_cell2.value = 'Kaupandi'
merged_cell2.font = bold_font
merged_cell2.alignment = Alignment(horizontal='center', vertical='center')

for cell in sheet["R"]:
    cell.number_format = thousand_separator_format
for cell in sheet["S"]:
    cell.number_format = thousand_separator_format
for cell in sheet["V"]:
    cell.number_format = thousand_separator_format

# Save the workbook to a file
workbook.save("C:\\Users\\sigurdurbl\\Desktop\\utskrift7.xlsx")

workbook.close()
